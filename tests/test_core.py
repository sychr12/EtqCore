from __future__ import annotations

import unittest
from contextlib import closing
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from app import app
from models import database
from models import etiqueta as etiqueta_model
from services.contador import counter_token
from services.impressao import recommended_zebra
from services import relatorio_excel
from services.texto import quantity_x1000
from services.validacao import validate_label, validate_settings


VALID_LABEL = {
    "tipo": "CAPA",
    "produto_codigo": "2000000444",
    "descricao": "Produto teste",
    "lote_controle": "BC",
    "quantidade": "95,5",
    "unidade": "pcs",
}


class CounterTests(unittest.TestCase):
    def test_numeric_and_alphanumeric_boundaries(self) -> None:
        self.assertEqual(counter_token(1), "0000000001")
        self.assertEqual(counter_token(9_999_999_999), "9999999999")
        self.assertEqual(counter_token(10_000_000_000), "A000000000")
        self.assertEqual(counter_token(10_999_999_999), "A999999999")
        self.assertEqual(counter_token(11_000_000_000), "B000000000")

    def test_counter_rejects_non_positive_values(self) -> None:
        with self.assertRaises(ValueError):
            counter_token(0)


class PrinterTests(unittest.TestCase):
    def test_zd220_is_preferred(self) -> None:
        printers = ["Microsoft Print to PDF", "ZDesigner GK420d", "ZDesigner ZD220-203dpi ZPL"]
        self.assertEqual(recommended_zebra(printers), "ZDesigner ZD220-203dpi ZPL")


class ValidationTests(unittest.TestCase):
    def test_quantity_formats(self) -> None:
        self.assertEqual(quantity_x1000("95,5"), "95500")
        self.assertEqual(quantity_x1000("1.234,5"), "1234500")

    def test_quantity_rejects_non_positive_and_non_finite(self) -> None:
        for value in ("0", "-1", "NaN", "Infinity"):
            with self.subTest(value=value), self.assertRaises(ValueError):
                quantity_x1000(value)

    def test_label_is_cleaned_and_validated(self) -> None:
        result = validate_label({**VALID_LABEL, "descricao": "Produto^\nTeste"})
        self.assertEqual(result["descricao"], "Produto  Teste")

    def test_tipo_is_optional(self) -> None:
        result = validate_label({**VALID_LABEL, "tipo": ""})
        self.assertEqual(result["tipo"], "")

    def test_text_is_normalized_for_utf8_zpl(self) -> None:
        result = validate_label({**VALID_LABEL, "descricao": "AÇA\u0303O"})
        self.assertEqual(result["descricao"], "AÇÃO")

    def test_observation_is_optional_and_normalized(self) -> None:
        result = validate_label({**VALID_LABEL, "observacao": "  REVISÃO ÇÃ  "})
        self.assertEqual(result["observacao"], "REVISÃO ÇÃ")

    def test_settings_reject_non_finite_dimensions(self) -> None:
        with self.assertRaises(ValueError):
                validate_settings({"largura_mm": "nan", "comprimento_mm": 60, "dpi": 203, "filial": "04"})

    def test_settings_enforce_zd220_profile(self) -> None:
        with self.assertRaises(ValueError):
            validate_settings({"largura_mm": 100, "comprimento_mm": 60, "dpi": 300, "filial": "04"})
        with self.assertRaises(ValueError):
            validate_settings({"largura_mm": 105, "comprimento_mm": 60, "dpi": 203, "filial": "04"})


class ApiTests(unittest.TestCase):
    def setUp(self) -> None:
        app.config.update(TESTING=True)
        self.client = app.test_client()

    def test_invalid_json_is_a_client_error(self) -> None:
        response = self.client.post("/api/gerar", data="not-json", content_type="application/json")
        self.assertEqual(response.status_code, 400)

    def test_invalid_destination_is_rejected_before_counter_reservation(self) -> None:
        response = self.client.post("/api/gerar", json={**VALID_LABEL, "destino": "arquivo-arbitrario"})
        self.assertEqual(response.status_code, 400)

    def test_qr_payload_has_a_size_limit(self) -> None:
        response = self.client.post("/api/qr.svg", data="x" * 4097)
        self.assertEqual(response.status_code, 400)

    def test_missing_printer_fails_before_reserving_counter(self) -> None:
        before = self.client.get("/api/estado").get_json()["proximo_numero"]
        with patch("services.geracao.resolve_printer", side_effect=RuntimeError("Zebra não encontrada")):
            response = self.client.post(
                "/api/gerar",
                json={**VALID_LABEL, "destino": "imprimir", "quantidade_etiquetas": 1},
            )
        after = self.client.get("/api/estado").get_json()["proximo_numero"]
        self.assertEqual(response.status_code, 409)
        self.assertFalse(response.get_json()["consumido"])
        self.assertEqual(after, before)

    def test_preview_reports_physical_print_dimensions(self) -> None:
        response = self.client.post("/api/preview", json=VALID_LABEL)
        self.assertEqual(response.status_code, 200)
        profile = response.get_json()["impressao"]
        self.assertEqual(profile["dpi"], 203)
        self.assertEqual(profile["largura_dots"], 800)
        self.assertEqual(profile["comprimento_dots"], 480)

    def test_preview_and_zpl_allow_empty_tipo(self) -> None:
        label = {**VALID_LABEL, "tipo": ""}
        response = self.client.post("/api/preview", json=label)
        self.assertEqual(response.status_code, 200)
        preview = response.get_json()
        self.assertIn("(T)(P)2000000444", preview["qr"])

        from services.zpl import make_zpl

        zpl = make_zpl(
            validate_label(label),
            1,
            "TB0000000001",
            preview["qr"],
            {
                "largura_mm": "100",
                "comprimento_mm": "60",
                "dpi": "203",
                "velocidade_ips": "3",
                "tonalidade": "10",
                "deslocamento_x_mm": "0",
                "deslocamento_y_mm": "0",
            },
        )
        self.assertIn("^XA", zpl)
        self.assertIn("^XZ", zpl)
        self.assertNotIn("^FDOBSERVAÇÃO:^FS", zpl)

    def test_observation_never_changes_qr_payload(self) -> None:
        from services.qrcode_service import qr_payload

        baseline = qr_payload(VALID_LABEL, "TB0000000001")
        for observation in ("", "   ", "Separar para inspecao"):
            self.assertEqual(qr_payload({**VALID_LABEL, "observacao": observation}, "TB0000000001"), baseline)
        self.assertNotIn("(O)", baseline)

    def test_black_brand_band_only_exists_when_client_is_filled(self) -> None:
        from services.qrcode_service import qr_payload
        from services.zpl import make_zpl

        settings = {
            "largura_mm": "100", "comprimento_mm": "60", "dpi": "203",
            "velocidade_ips": "3", "tonalidade": "10",
            "deslocamento_x_mm": "0", "deslocamento_y_mm": "0",
        }
        without_client = validate_label({**VALID_LABEL, "cliente": ""})
        zpl_without = make_zpl(without_client, 1, "TB0000000001", qr_payload(without_client, "TB0000000001"), settings)
        self.assertNotIn("^FO700,28^GB72,424,72,B,0^FS", zpl_without)
        self.assertIn("^FO28,232^GB744,4,4,B,0^FS", zpl_without)

        with_client = validate_label({**VALID_LABEL, "cliente": "AMAZONTAPE", "lote_base": "4016/2026", "quantidade": "910"})
        zpl_with = make_zpl(with_client, 1, "TB0000000001", qr_payload(with_client, "TB0000000001"), settings)
        self.assertIn("^FO700,28^GB72,424,72,B,0^FS", zpl_with)
        self.assertIn("^FDAMAZONTAPE^FS", zpl_with)
        self.assertIn("^FD2026/4016^FS", zpl_with)
        self.assertIn("^FD910 pcs^FS", zpl_with)

    def test_windows_folder_picker_returns_selected_path(self) -> None:
        selected = r"C:\Relatorios Compartilhados"
        with patch("controllers.etiquetas_controller.escolher_pasta_windows", return_value=selected):
            response = self.client.post(
                "/api/relatorios/pasta/escolher",
                json={"pasta_atual": r"C:\Relatorios"},
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["pasta"], selected)
        self.assertFalse(response.get_json()["cancelado"])

    def test_selected_report_folder_is_saved_without_changing_printer_settings(self) -> None:
        selected = Path(r"C:\Relatorios Compartilhados")
        with patch("controllers.etiquetas_controller.testar_pasta", return_value=selected), \
             patch("controllers.etiquetas_controller.config_model.salvar") as save:
            response = self.client.post("/api/relatorios/pasta/configurar", json={"pasta": str(selected)})
        self.assertEqual(response.status_code, 200)
        save.assert_called_once_with({"pasta_relatorios": str(selected)})

    def test_monthly_excel_is_saved_inside_year_and_month_folders(self) -> None:
        with TemporaryDirectory() as directory:
            old_reports = relatorio_excel.REPORTS_DIR
            try:
                relatorio_excel.REPORTS_DIR = Path(directory)
                arquivo = relatorio_excel.criar_relatorio_mensal(2026, 8, [{
                    "id": 1,
                    "contador": 1,
                    "identificador": "TB0000000001",
                    "criada_em": "2026-08-05T10:30:00",
                    "dados": {**VALID_LABEL, "descricao": "AÇÃO ÇÃ", "operador": "252"},
                    "destino": "imprimir",
                    "sucesso": 1,
                    "erro": None,
                }])
                self.assertEqual(arquivo, Path(directory) / "2026" / "08" / f"etiquetas-2026-08-{datetime.now().day:02d}.xlsx")
                self.assertTrue(arquivo.exists())

                from openpyxl import load_workbook
                workbook = load_workbook(arquivo, read_only=True)
                sheet = workbook["Etiquetas"]
                self.assertEqual(sheet["A1"].value, "RELATÓRIO DE ETIQUETAS — AGOSTO DE 2026")
                self.assertEqual(sheet["C7"].value, "TB0000000001")
                self.assertEqual(sheet["H7"].value, "AÇÃO ÇÃ")
                workbook.close()
            finally:
                relatorio_excel.REPORTS_DIR = old_reports

    def test_annual_excel_has_summary_and_month_sheets(self) -> None:
        with TemporaryDirectory() as directory:
            old_reports = relatorio_excel.REPORTS_DIR
            try:
                relatorio_excel.REPORTS_DIR = Path(directory)
                base = {
                    "id": 1, "contador": 1, "identificador": "TB0000000001",
                    "dados": VALID_LABEL, "destino": "imprimir", "sucesso": 1, "erro": None,
                }
                arquivo = relatorio_excel.criar_relatorio_anual(2026, [
                    {**base, "criada_em": "2026-01-10T08:00:00"},
                    {**base, "id": 2, "contador": 2, "identificador": "TB0000000002", "criada_em": "2026-08-05T10:30:00", "sucesso": 0, "erro": "Teste"},
                ])
                self.assertEqual(arquivo, Path(directory) / "2026" / "etiquetas-2026-anual.xlsx")

                from openpyxl import load_workbook
                workbook = load_workbook(arquivo, read_only=True, data_only=False)
                self.assertIn("Resumo anual", workbook.sheetnames)
                self.assertIn("01 - Janeiro", workbook.sheetnames)
                self.assertIn("08 - Agosto", workbook.sheetnames)
                self.assertNotIn("02 - Fevereiro", workbook.sheetnames)
                self.assertEqual(workbook["Resumo anual"]["B16"].value, "=SUM(B4:B15)")
                workbook.close()
            finally:
                relatorio_excel.REPORTS_DIR = old_reports

    def test_reports_folder_can_be_external_and_writable(self) -> None:
        with TemporaryDirectory() as directory:
            external = Path(directory) / "Compartilhado" / "Etiquetas"
            resolved = relatorio_excel.testar_pasta(external)
            self.assertTrue(resolved.is_dir())
            self.assertFalse((resolved / ".etqcore-teste.tmp").exists())

            file_path = relatorio_excel.caminho_relatorio(2026, 8, external)
            self.assertEqual(file_path, external / "2026" / "08" / f"etiquetas-2026-08-{datetime.now().day:02d}.xlsx")

    def test_generation_reserves_unique_counters_and_returns_zpl(self) -> None:
        with TemporaryDirectory() as directory:
            old_data_dir, old_db_path, old_backup_dir = (
                database.DATA_DIR,
                database.DB_PATH,
                database.BACKUP_DIR,
            )
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "test.db"
                database.BACKUP_DIR = Path(directory) / "backups"
                database.init_db()

                response = self.client.post(
                    "/api/gerar",
                    json={
                        **VALID_LABEL,
                        "cliente": "AMAZONTAPE",
                        "descricao": "TUBETE AÇÃO ÇÃ",
                        "observacao": "Separar para inspeção",
                        "destino": "download",
                        "quantidade_etiquetas": 2,
                    },
                )
                payload = response.get_json()
                self.assertEqual(response.status_code, 200)
                self.assertEqual(payload["quantidade"], 2)
                self.assertEqual(payload["identificador"], "TB0000000001")
                self.assertEqual(payload["ultimo_identificador"], "TB0000000002")
                self.assertEqual(payload["zpl"].count("^XA"), 2)
                self.assertIn("^CI28", payload["zpl"])
                self.assertNotIn("^CI27", payload["zpl"])
                self.assertIn("^MNA", payload["zpl"])
                self.assertIn("^PR3", payload["zpl"])
                self.assertIn("~SD10", payload["zpl"])
                self.assertIn("^PW800", payload["zpl"])
                self.assertIn("^LL480", payload["zpl"])
                self.assertIn("^BQN,2,4", payload["zpl"])
                self.assertIn("TUBETE AÇÃO ÇÃ", payload["zpl"])
                self.assertIn("^GFA", payload["zpl"])
                # Tipo e lote de controle permanecem no QR, mas não são
                # desenhados como textos/blocos visuais na etiqueta.
                self.assertIn("(T)CAPA", payload["zpl"])
                self.assertIn("(S)BC", payload["zpl"])
                self.assertNotIn("(O)Separar para inspeção", payload["zpl"])
                self.assertIn("^FDSeparar para inspeção^FS", payload["zpl"])
                self.assertIn("^FDOBSERVAÇÃO:^FS", payload["zpl"])
                self.assertIn("Separar para inspeção", payload["zpl"])
                self.assertNotIn("^FO232,28^GB88,204,88,B,0^FS", payload["zpl"])
                self.assertNotIn("^FDCAPA^FS", payload["zpl"])
                self.assertNotIn("^FDBC^FS", payload["zpl"])
                # Separadores das quatro faixas do novo layout com QR maior.
                self.assertIn("^FO28,232^GB672,4,4,B,0^FS", payload["zpl"])
                self.assertIn("^FO28,292^GB672,4,4,B,0^FS", payload["zpl"])
                self.assertIn("^FO28,368^GB672,4,4,B,0^FS", payload["zpl"])

                with closing(database.db()) as connection:
                    next_counter = connection.execute(
                        "SELECT valor FROM config WHERE chave='proximo_contador'"
                    ).fetchone()[0]
                    label_count = connection.execute("SELECT COUNT(*) FROM etiquetas").fetchone()[0]
                self.assertEqual(next_counter, "3")
                self.assertEqual(label_count, 2)
            finally:
                database.DATA_DIR = old_data_dir
                database.DB_PATH = old_db_path
                database.BACKUP_DIR = old_backup_dir

    def test_undo_latest_label_restores_counter_and_creates_backup(self) -> None:
        with TemporaryDirectory() as directory:
            old_data, old_db, old_database_backup, old_label_backup = (
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR,
            )
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "undo.db"
                database.BACKUP_DIR = Path(directory) / "backups"
                etiqueta_model.BACKUP_DIR = database.BACKUP_DIR
                database.init_db()
                with closing(database.db()) as con:
                    con.execute(
                        "INSERT INTO etiquetas(contador,identificador,criada_em,dados_json,qr_texto,zpl,destino,sucesso) VALUES(1,'TB0000000001','2026-09-03T10:00:00','{}','QR','ZPL','download',1)"
                    )
                    con.execute("UPDATE config SET valor='2' WHERE chave='proximo_contador'")
                    label_id = con.execute("SELECT id FROM etiquetas").fetchone()[0]
                    con.commit()

                response = self.client.post(f"/api/historico/{label_id}/desfazer")
                self.assertEqual(response.status_code, 200)
                with closing(database.db()) as con:
                    self.assertEqual(con.execute("SELECT COUNT(*) FROM etiquetas").fetchone()[0], 0)
                    self.assertEqual(con.execute("SELECT valor FROM config WHERE chave='proximo_contador'").fetchone()[0], "1")
                self.assertTrue(Path(response.get_json()["backup"]).exists())
            finally:
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR = (
                    old_data, old_db, old_database_backup, old_label_backup,
                )

    def test_undo_rejects_an_older_label(self) -> None:
        with TemporaryDirectory() as directory:
            old_data, old_db = database.DATA_DIR, database.DB_PATH
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "undo-guard.db"
                database.init_db()
                with closing(database.db()) as con:
                    for contador in (1, 2):
                        con.execute(
                            "INSERT INTO etiquetas(contador,identificador,criada_em,dados_json,qr_texto,zpl,destino,sucesso) VALUES(?,?,?,?,?,?,?,1)",
                            (contador, f"TB{contador:010d}", "2026-09-03T10:00:00", "{}", "QR", "ZPL", "download"),
                        )
                    con.execute("UPDATE config SET valor='3' WHERE chave='proximo_contador'")
                    older_id = con.execute("SELECT id FROM etiquetas WHERE contador=1").fetchone()[0]
                    con.commit()
                response = self.client.post(f"/api/historico/{older_id}/desfazer")
                self.assertEqual(response.status_code, 409)
                with closing(database.db()) as con:
                    self.assertEqual(con.execute("SELECT COUNT(*) FROM etiquetas").fetchone()[0], 2)
                    self.assertEqual(con.execute("SELECT valor FROM config WHERE chave='proximo_contador'").fetchone()[0], "3")
            finally:
                database.DATA_DIR, database.DB_PATH = old_data, old_db

    def test_clear_history_creates_backup_deletes_rows_and_resets_counter(self) -> None:
        with TemporaryDirectory() as directory:
            old_data, old_db, old_database_backup, old_label_backup = (
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR,
            )
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "clear-history.db"
                database.BACKUP_DIR = Path(directory) / "backups"
                etiqueta_model.BACKUP_DIR = database.BACKUP_DIR
                database.init_db()
                with closing(database.db()) as con:
                    for contador in (1, 2, 3):
                        con.execute(
                            "INSERT INTO etiquetas(contador,identificador,criada_em,dados_json,qr_texto,zpl,destino,sucesso) VALUES(?,?,?,?,?,?,?,1)",
                            (contador, f"TB{contador:010d}", "2026-09-03T10:00:00", "{}", "QR", "ZPL", "download"),
                        )
                    con.execute("UPDATE config SET valor='4' WHERE chave='proximo_contador'")
                    con.commit()

                response = self.client.post("/api/historico/apagar-tudo")
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.get_json()["registros_apagados"], 3)
                with closing(database.db()) as con:
                    self.assertEqual(con.execute("SELECT COUNT(*) FROM etiquetas").fetchone()[0], 0)
                    self.assertEqual(con.execute("SELECT valor FROM config WHERE chave='proximo_contador'").fetchone()[0], "1")
                self.assertTrue(Path(response.get_json()["backup"]).exists())
            finally:
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR = (
                    old_data, old_db, old_database_backup, old_label_backup,
                )

    def test_clear_history_range_sets_exact_next_counter_and_creates_backup(self) -> None:
        with TemporaryDirectory() as directory:
            old_data, old_db, old_database_backup, old_label_backup = (
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR,
            )
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "clear-range.db"
                database.BACKUP_DIR = Path(directory) / "backups"
                etiqueta_model.BACKUP_DIR = database.BACKUP_DIR
                database.init_db()
                with closing(database.db()) as con:
                    for contador in (1, 2, 3):
                        con.execute(
                            "INSERT INTO etiquetas(contador,identificador,criada_em,dados_json,qr_texto,zpl,destino,sucesso) VALUES(?,?,?,?,?,?,?,1)",
                            (contador, f"TB{contador:010d}", "2026-09-03T10:00:00", "{}", "QR", "ZPL", "download"),
                        )
                    con.execute("UPDATE config SET valor='4' WHERE chave='proximo_contador'")
                    con.commit()

                response = self.client.post(
                    "/api/historico/apagar-intervalo",
                    json={"inicio": 1, "fim": 3, "proximo": 89},
                )
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.get_json()["registros_apagados"], 3)
                with closing(database.db()) as con:
                    self.assertEqual(con.execute("SELECT COUNT(*) FROM etiquetas").fetchone()[0], 0)
                    self.assertEqual(con.execute("SELECT valor FROM config WHERE chave='proximo_contador'").fetchone()[0], "89")
                self.assertTrue(Path(response.get_json()["backup"]).exists())
            finally:
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR = (
                    old_data, old_db, old_database_backup, old_label_backup,
                )

    def test_delete_individual_labels_preserves_counter_and_other_rows(self) -> None:
        with TemporaryDirectory() as directory:
            old_data, old_db, old_database_backup, old_label_backup = (
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR,
            )
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "clear-range.db"
                database.BACKUP_DIR = Path(directory) / "backups"
                etiqueta_model.BACKUP_DIR = database.BACKUP_DIR
                database.init_db()
                with closing(database.db()) as con:
                    for contador in (1, 2, 3):
                        con.execute(
                            "INSERT INTO etiquetas(contador,identificador,criada_em,dados_json,qr_texto,zpl,destino,sucesso) VALUES(?,?,?,?,?,?,?,1)",
                            (contador, f"TB{contador:010d}", "2026-09-03T10:00:00", "{}", "QR", "ZPL", "download"),
                        )
                    con.execute("UPDATE config SET valor='4' WHERE chave='proximo_contador'")
                    con.commit()

                with closing(database.db()) as con:
                    ids = {row["contador"]: row["id"] for row in con.execute("SELECT id, contador FROM etiquetas")}
                for counter in (2, 3):
                    response = self.client.delete(f"/api/historico/{ids[counter]}")
                    self.assertEqual(response.status_code, 200)
                    self.assertTrue(Path(response.get_json()["backup"]).exists())
                    with closing(database.db()) as con:
                        self.assertEqual(con.execute("SELECT valor FROM config WHERE chave='proximo_contador'").fetchone()[0], "4")
                with closing(database.db()) as con:
                    self.assertEqual([row[0] for row in con.execute("SELECT contador FROM etiquetas")], [1])
                self.assertEqual(self.client.delete(f"/api/historico/{ids[2]}").status_code, 404)
            finally:
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR = (
                    old_data, old_db, old_database_backup, old_label_backup,
                )

    def test_delete_last_twenty_preserves_first_eighty_and_restores_81(self) -> None:
        with TemporaryDirectory() as directory:
            old_data, old_db, old_database_backup, old_label_backup = (
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR,
            )
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "clear-range.db"
                database.BACKUP_DIR = Path(directory) / "backups"
                etiqueta_model.BACKUP_DIR = database.BACKUP_DIR
                database.init_db()
                with closing(database.db()) as con:
                    for contador in range(1, 101):
                        con.execute(
                            "INSERT INTO etiquetas(contador,identificador,criada_em,dados_json,qr_texto,zpl,destino,sucesso) VALUES(?,?,?,?,?,?,?,1)",
                            (contador, f"TB{contador:010d}", "2026-09-03T10:00:00", "{}", "QR", "ZPL", "download"),
                        )
                    con.execute("UPDATE config SET valor='101' WHERE chave='proximo_contador'")
                    con.commit()

                response = self.client.post(
                    "/api/historico/apagar-intervalo",
                    json={"inicio": 81, "fim": 100, "proximo": 81},
                )
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.get_json()["registros_apagados"], 20)
                with closing(database.db()) as con:
                    self.assertEqual(con.execute("SELECT COUNT(*) FROM etiquetas").fetchone()[0], 80)
                    self.assertEqual(con.execute("SELECT valor FROM config WHERE chave='proximo_contador'").fetchone()[0], "81")
                    self.assertEqual([row[0] for row in con.execute("SELECT contador FROM etiquetas ORDER BY contador")], list(range(1, 81)))
                self.assertTrue(Path(response.get_json()["backup"]).exists())
            finally:
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR = (
                    old_data, old_db, old_database_backup, old_label_backup,
                )

    def test_delete_top_rows_counts_records_despite_gaps(self) -> None:
        with TemporaryDirectory() as directory:
            old_data, old_db, old_database_backup, old_label_backup = (
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR,
            )
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "clear-range.db"
                database.BACKUP_DIR = Path(directory) / "backups"
                etiqueta_model.BACKUP_DIR = database.BACKUP_DIR
                database.init_db()
                with closing(database.db()) as con:
                    for contador in (1, 2, 4, 5):
                        con.execute(
                            "INSERT INTO etiquetas(contador,identificador,criada_em,dados_json,qr_texto,zpl,destino,sucesso) VALUES(?,?,?,?,?,?,?,1)",
                            (contador, f"TB{contador:010d}", "2026-09-03T10:00:00", "{}", "QR", "ZPL", "download"),
                        )
                    con.execute("UPDATE config SET valor='6' WHERE chave='proximo_contador'")
                    con.commit()

                preview = self.client.post("/api/historico/apagar-ultimas/previa", json={"quantidade": 2}).get_json()
                self.assertEqual((preview["primeiro"], preview["ultimo"], preview["ultimo_correto"]), (5, 4, 2))
                stale = self.client.post("/api/historico/apagar-ultimas", json={"quantidade": 2, "ids": [0, 1]})
                self.assertEqual(stale.status_code, 400)
                response = self.client.post("/api/historico/apagar-ultimas", json={"quantidade": 2, "ids": preview["ids"]})
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.get_json()["registros_apagados"], 2)
                self.assertEqual(response.get_json()["proximo_contador"], 3)
                self.assertTrue(Path(response.get_json()["backup"]).exists())
                with closing(database.db()) as con:
                    self.assertEqual([r[0] for r in con.execute("SELECT contador FROM etiquetas ORDER BY id")], [1, 2])
            finally:
                database.DATA_DIR, database.DB_PATH, database.BACKUP_DIR, etiqueta_model.BACKUP_DIR = (
                    old_data, old_db, old_database_backup, old_label_backup,
                )

    def test_clear_history_range_rejects_counter_conflict(self) -> None:
        with TemporaryDirectory() as directory:
            old_data, old_db = database.DATA_DIR, database.DB_PATH
            try:
                database.DATA_DIR = Path(directory)
                database.DB_PATH = Path(directory) / "range-conflict.db"
                database.init_db()
                with closing(database.db()) as con:
                    for contador in (1, 2, 3, 4):
                        con.execute(
                            "INSERT INTO etiquetas(contador,identificador,criada_em,dados_json,qr_texto,zpl,destino,sucesso) VALUES(?,?,?,?,?,?,?,1)",
                            (contador, f"TB{contador:010d}", "2026-09-03T10:00:00", "{}", "QR", "ZPL", "download"),
                        )
                    con.commit()
                response = self.client.post(
                    "/api/historico/apagar-intervalo",
                    json={"inicio": 1, "fim": 3, "proximo": 2},
                )
                self.assertEqual(response.status_code, 400)
                with closing(database.db()) as con:
                    self.assertEqual(con.execute("SELECT COUNT(*) FROM etiquetas").fetchone()[0], 4)
            finally:
                database.DATA_DIR, database.DB_PATH = old_data, old_db


if __name__ == "__main__":
    unittest.main()
