"""Cria planilhas mensais e anuais do histórico de etiquetas."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR


MESES = (
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
)


def caminho_relatorio(ano: int, mes: int, pasta_base: str | Path | None = None) -> Path:
    """Retorna um caminho previsível e seguro para o arquivo mensal."""
    if not 2000 <= ano <= 2100 or not 1 <= mes <= 12:
        raise ValueError("Ano ou mês inválido.")
    base = Path(pasta_base) if pasta_base else REPORTS_DIR
    return base / str(ano) / f"{mes:02d}" / f"etiquetas-{ano:04d}-{mes:02d}.xlsx"


def caminho_relatorio_anual(ano: int, pasta_base: str | Path | None = None) -> Path:
    """Retorna o arquivo consolidado que fica diretamente na pasta do ano."""
    if not 2000 <= ano <= 2100:
        raise ValueError("Ano inválido.")
    base = Path(pasta_base) if pasta_base else REPORTS_DIR
    return base / str(ano) / f"etiquetas-{ano:04d}-anual.xlsx"


def testar_pasta(pasta: str | Path) -> Path:
    """Confirma que a pasta local ou de rede existe e aceita gravação."""
    destino = Path(pasta).expanduser()
    destino.mkdir(parents=True, exist_ok=True)
    teste = destino / ".etqcore-teste.tmp"
    try:
        teste.write_text("teste", encoding="utf-8")
        teste.unlink()
    except OSError as exc:
        raise ValueError(f"Não foi possível gravar na pasta: {exc}") from exc
    return destino.resolve()


def escolher_pasta_windows(pasta_inicial: str | Path | None = None) -> str | None:
    """Abre o seletor nativo de pastas no computador onde o programa roda."""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError as exc:
        raise RuntimeError("O seletor de pastas do Windows não está disponível.") from exc

    inicial = Path(pasta_inicial).expanduser() if pasta_inicial else REPORTS_DIR
    if not inicial.exists():
        inicial = Path.home() / "Documents"
    janela = tk.Tk()
    janela.withdraw()
    janela.attributes("-topmost", True)
    try:
        escolha = filedialog.askdirectory(
            parent=janela,
            initialdir=str(inicial),
            title="Escolha onde salvar os relatórios de etiquetas",
            mustexist=False,
        )
        return escolha or None
    finally:
        janela.destroy()


HEADERS = [
    "ID", "Contador", "Identificador", "Data e hora", "Tipo", "Produto",
    "Código interno", "Descrição", "Lote controle", "Lote fabricação",
    "Quantidade", "Unidade", "Operador", "Medidas", "Destino", "Status", "Erro",
]


def _preencher_planilha(sheet, titulo: str, etiquetas: list[dict]) -> None:
    """Aplica o mesmo padrão visual e os mesmos campos em qualquer aba."""
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A7"
    sheet.merge_cells("A1:Q1")
    sheet["A1"] = titulo
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0B6B5C")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 34

    indicadores = (("A3", "Total", len(etiquetas)),
                    ("D3", "Sucesso", sum(1 for item in etiquetas if item["sucesso"])),
                    ("G3", "Falhas", sum(1 for item in etiquetas if not item["sucesso"])))
    for celula, rotulo, valor in indicadores:
        sheet[celula] = rotulo
        sheet[celula].font = Font(bold=True, color="40566A")
        valor_cell = sheet.cell(row=3, column=sheet[celula].column + 1, value=valor)
        valor_cell.font = Font(size=15, bold=True, color="0B6B5C")
        valor_cell.number_format = "#,##0"
    sheet["A4"] = "Gerado em"
    sheet["B4"] = datetime.now()
    sheet["B4"].number_format = "dd/mm/yyyy hh:mm"

    sheet.append([])
    sheet.append(HEADERS)
    header_row = 6
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="173B56")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 30

    for item in etiquetas:
        dados = item["dados"]
        try:
            data_hora = datetime.fromisoformat(item["criada_em"])
        except ValueError:
            data_hora = item["criada_em"]
        sheet.append([
            item["id"], item["contador"], item["identificador"], data_hora,
            dados.get("tipo", ""), dados.get("produto_codigo", ""), dados.get("cod_prod", ""),
            dados.get("descricao", ""), dados.get("lote_controle", ""), dados.get("lote_base", ""),
            dados.get("quantidade", ""), dados.get("unidade", ""), dados.get("operador", ""),
            dados.get("medidas", ""), item["destino"], "Sucesso" if item["sucesso"] else "Falha",
            item["erro"] or "",
        ])

    final_row = max(header_row, sheet.max_row)
    sheet.auto_filter.ref = f"A{header_row}:Q{final_row}"
    thin = Side(style="thin", color="D9E2E8")
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=final_row):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in {8, 17})
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F2F7F6")
    for row in range(header_row + 1, final_row + 1):
        sheet.cell(row=row, column=4).number_format = "dd/mm/yyyy hh:mm:ss"
        status_cell = sheet.cell(row=row, column=16)
        status_cell.font = Font(bold=True, color="08735B" if status_cell.value == "Sucesso" else "B42318")

    widths = [8, 12, 18, 20, 12, 18, 18, 34, 16, 18, 14, 11, 14, 24, 13, 12, 38]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def criar_relatorio_mensal(ano: int, mes: int, etiquetas: list[dict], pasta_base: str | Path | None = None) -> Path:
    """Gera uma planilha formatada contendo todas as etiquetas do período."""
    destino = caminho_relatorio(ano, mes, pasta_base)
    destino.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Etiquetas"
    _preencher_planilha(sheet, f"RELATÓRIO DE ETIQUETAS — {MESES[mes - 1].upper()} DE {ano}", etiquetas)

    workbook.save(destino)
    return destino


def criar_relatorio_anual(ano: int, etiquetas: list[dict], pasta_base: str | Path | None = None) -> Path:
    """Cria um consolidado anual com resumo e uma aba para cada mês com dados."""
    destino = caminho_relatorio_anual(ano, pasta_base)
    destino.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    resumo = workbook.active
    resumo.title = "Resumo anual"
    resumo.sheet_view.showGridLines = False
    resumo.merge_cells("A1:D1")
    resumo["A1"] = f"RESUMO ANUAL DE ETIQUETAS — {ano}"
    resumo["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    resumo["A1"].fill = PatternFill("solid", fgColor="0B6B5C")
    resumo["A1"].alignment = Alignment(horizontal="center")
    resumo.row_dimensions[1].height = 34
    resumo.append([])
    resumo.append(["Mês", "Total", "Sucesso", "Falhas"])
    for cell in resumo[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="173B56")

    por_mes = {mes: [] for mes in range(1, 13)}
    for item in etiquetas:
        try:
            mes = datetime.fromisoformat(item["criada_em"]).month
        except ValueError:
            continue
        por_mes[mes].append(item)
    for mes in range(1, 13):
        itens = por_mes[mes]
        resumo.append([MESES[mes - 1], len(itens), sum(1 for i in itens if i["sucesso"]), sum(1 for i in itens if not i["sucesso"])])
        if itens:
            aba = workbook.create_sheet(f"{mes:02d} - {MESES[mes - 1]}")
            _preencher_planilha(aba, f"ETIQUETAS — {MESES[mes - 1].upper()} DE {ano}", itens)
    total_row = 16
    resumo.cell(total_row, 1, "TOTAL DO ANO").font = Font(bold=True, color="FFFFFF")
    for col in range(1, 5):
        resumo.cell(total_row, col).fill = PatternFill("solid", fgColor="0B6B5C")
    resumo.cell(total_row, 2, f"=SUM(B4:B15)")
    resumo.cell(total_row, 3, f"=SUM(C4:C15)")
    resumo.cell(total_row, 4, f"=SUM(D4:D15)")
    for col in range(2, 5):
        resumo.cell(total_row, col).font = Font(bold=True, color="FFFFFF")
        resumo.cell(total_row, col).number_format = "#,##0"
    resumo.freeze_panes = "A4"
    resumo.auto_filter.ref = "A3:D15"
    resumo.column_dimensions["A"].width = 20
    for col in ("B", "C", "D"):
        resumo.column_dimensions[col].width = 15
    workbook.save(destino)
    return destino
